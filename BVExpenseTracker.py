import openpyxl
import os

def main():
    # Insert the file of the spreadsheet you want to change here
    fileName = 'November2019Transactions.xlsx'

    # Make sure to go into the Excel sheet and change the sheet name to transactions prior to running the program
    # Sheet name you would like to change
    sheetName = 'transactions'
    # Load the workbook
    wb = openpyxl.load_workbook(fileName)

    # Create a new sheet
    createResultsSheet(wb, fileName)
    
    # Access the sheet in the workbook
    transactionsWorksheet = wb[sheetName]
    resultsWorksheet = wb['results']

    # If E1 is tags, then delete column 5
    delete_tags_column(transactionsWorksheet)
    wb.save(fileName)

    # Fix double payments
    fixPayments(transactionsWorksheet)
    wb.save(fileName)

    # Get all the transactions in categories
    organizeExpenses(transactionsWorksheet, resultsWorksheet)
    wb.save(fileName)




def delete_tags_column(transactionsWorksheet):
    tagsValue = transactionsWorksheet['E1'].value
    if (tagsValue == 'Tags'):
        transactionsWorksheet.delete_cols(5)

def organizeExpenses(transactionsWorksheet, resultsWorksheet):
    # Maximum row of the sheet
    maxRow = transactionsWorksheet.max_row

    # Declare empty dictionary to store category as key  and total expense as value
    transactionCategories = {}

    
    # Iterate over each row to obtain category and transition
    for i in range(2, maxRow + 1):
        category = transactionsWorksheet.cell(row=i, column=4).value
        transaction = transactionsWorksheet.cell(row = i, column=5).value

        # If category exists, add the transaction to the total expense
        if category in transactionCategories:
            transactionCategories[category] += transaction
        # If category does not, make a new one and set it to the first transaction that appears
        elif category not in transactionCategories:
            transactionCategories[category] = transaction
    
    categories = list(transactionCategories.keys())

    for i in range(2, len(categories) + 2):
        cellValue = ('A' + str(i))
        resultsWorksheet[cellValue] = categories[i - 2]
        expenseCell = ('B' + str(i))
        resultsWorksheet[expenseCell] = transactionCategories[categories[i-2]]

    resultsWorksheet['A1'] = 'Category'
    resultsWorksheet['B1'] = 'Amount'

    for key in transactionCategories:
        print(key, '$', round(transactionCategories[key],2))
    
def fixPayments(incomeExpensesSheet):
    maxRow = incomeExpensesSheet.max_row

    for i in range(2, maxRow + 1):
        if (str(incomeExpensesSheet.cell(row=i, column= 4).value) == "Credit Card Payments") or (str(incomeExpensesSheet.cell(row=i, column= 3).value) == "Transfer" and str(incomeExpensesSheet.cell(row=i, column= 2).value) == "Trs Plan 3 - Self") or (incomeExpensesSheet.cell(row=i, column= 3).value == "Reinvestment Fidelity 500 Index Fund"):
            cellValue = ('E' + str(i))
            incomeExpensesSheet[cellValue] = 0
            transactionValue = 0

        if(len(str(incomeExpensesSheet.cell(row=i, column=3).value)) > 14):
            if str(incomeExpensesSheet.cell(row=i, column=3).value)[0:14] == "Paccar Kenwort":
                cellValue = ('D' + str(i))
                incomeExpensesSheet[cellValue] = "Restaurants"
                cellValue = ('C' + str(i))
                incomeExpensesSheet[cellValue] = "Kenworth Cafeteria"

            elif str(incomeExpensesSheet.cell(row=i, column=3).value)[0:7] == "Hunt-bw":
                cellValue = ('D' + str(i))
                incomeExpensesSheet[cellValue] = "Rent"
                cellValue = ('C' + str(i))
                incomeExpensesSheet[cellValue] = "Bridlwood Apartments"

        #Updating Gold's Gym info
        if(str(incomeExpensesSheet.cell(row=i, column=3).value) == "4610 Gg Kirkland Kirkland Wa"):
            cellValue = ('D' + str(i))
            incomeExpensesSheet[cellValue] = "Fitness"
            cellValue = ('C' + str(i))
            incomeExpensesSheet[cellValue] = "Gold's Gym"

def createResultsSheet(wb, fileName):
    sheets = wb.sheetnames
    if 'results' in sheets:
        print('found results')
    else:
        wb.create_sheet('results')
        print('made a new sheet called results')
        wb.save(fileName)
        

if __name__ == "__main__":
    main()

