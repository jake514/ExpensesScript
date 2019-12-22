#steps: 
# 1) User enters date range for data collection
# 2) Python pulls info for that date range from excel file
# 3) Python displays sum of positive values(income), sum of negative values(expenses), savings rate (income-expenses/income) * 100
# 4) show breakdown of income sources (Betty's job)
import os
import re
import openpyxl

# Loads expenses workbook from personal capital
wb = openpyxl.load_workbook('expenses.xlsx')
incomeExpensesSheet = wb['Income_Expenses']
calculationsSheet = wb['Calculations']

#variable initiation
income=0
expenses=0

#Finds last row value and earliest date
lastRow = incomeExpensesSheet.max_row
i = lastRow
firstDate = str(incomeExpensesSheet.cell(row=lastRow, column=1).value)
firstDate = firstDate[5:7]

#initialize dictionary and first stored array
monthExpenses = {}
monthExpenses[firstDate] = []


while i>1:

    #find month of cell and compare to previous cell. If different, add another value to dictionary.
    cellMonth = str(incomeExpensesSheet.cell(row=i, column=1).value)
    cellMonth = cellMonth[5:7]
    lastCellMonth = str(incomeExpensesSheet.cell(row=i+1, column=1).value)
    lastCellMonth = lastCellMonth[5:7]

    if cellMonth != lastCellMonth and i != lastRow:
        
        #Adds all calculations from previous month to dictionary
        expenses *= -1
        income = round(income, 2)
        expenses = round(expenses,2)
        if income != 0:
            savingsRate = round((income - expenses) / income * 100, 2)
        else:
            savingsRate = 0
        netIncome = round(income - expenses,2)
        monthExpenses[lastCellMonth] = []
        monthExpenses[lastCellMonth].append(income)
        monthExpenses[lastCellMonth].append(expenses)
        monthExpenses[lastCellMonth].append(savingsRate)
        monthExpenses[lastCellMonth].append(netIncome)

        #zeroes out the values for next month calculations
        income = 0
        expenses = 0

        
    transactionValue = incomeExpensesSheet.cell(row=i, column=5).value
        
    #corrects transaction values to 0 if not required
    if (str(incomeExpensesSheet.cell(row=i, column= 4).value) == "Credit Card Payments") or (str(incomeExpensesSheet.cell(row=i, column= 3).value) == "Transfer" and str(incomeExpensesSheet.cell(row=i, column= 2).value) == "Trs Plan 3 - Self") or (incomeExpensesSheet.cell(row=i, column= 3).value == "Reinvestment Fidelity 500 Index Fund"):
        cellValue = ('E' + str(i))
        incomeExpensesSheet[cellValue] = 0
        transactionValue = 0

    #corrects payment category for purchases at kenworth
    if(len(str(incomeExpensesSheet.cell(row=i, column=3).value)) > 14):
        if str(incomeExpensesSheet.cell(row=i, column=3).value)[0:14] == "Paccar Kenwort":
            cellValue = ('D' + str(i))
            incomeExpensesSheet[cellValue] = "Restaurants"
            cellValue = ('C' + str(i))
            incomeExpensesSheet[cellValue] = "Kenworth Cafeteria"

        elif str(incomeExpensesSheet.cell(row=i, column=3).value)[0:7] == "Hunt-bw":
            cellValue = ('D' + str(i))
            incomeExpensesSheet[cellValue] = "Rent"
            cellValue = ('C' + str(i))
            incomeExpensesSheet[cellValue] = "Bridlwood Apartments"
    
    #Updating Gold's Gym info
    if(str(incomeExpensesSheet.cell(row=i, column=3).value) == "4610 Gg Kirkland Kirkland Wa"):
        cellValue = ('D' + str(i))
        incomeExpensesSheet[cellValue] = "Fitness"
        cellValue = ('C' + str(i))
        incomeExpensesSheet[cellValue] = "Gold's Gym"    

    if transactionValue >= 0:
        income += transactionValue
    else:
        expenses += transactionValue

    i -= 1

# --------------------- Print Results to sheet 2 ------------------------------
j=2

for month in monthExpenses.keys():
    calculationsSheet['A1'] = "Month"
    calculationsSheet['B1'] = "Income"
    calculationsSheet['C1'] = "Expenses"
    calculationsSheet['D1'] = "Savings Rate"
    calculationsSheet['E1'] = "Net Income"

    
    cellValue = ('A' + str(j))
    calculationsSheet[cellValue] = month

    cellValue = ('B' + str(j))
    calculationsSheet[cellValue] = monthExpenses[month][0]

    cellValue = ('C' + str(j))
    calculationsSheet[cellValue] = monthExpenses[month][1]

    
    cellValue = ('D' + str(j))
    calculationsSheet[cellValue] = monthExpenses[month][2]

    cellValue = ('E' + str(j))
    calculationsSheet[cellValue] = monthExpenses[month][3]
    j += 2

wb.save('expenses.xlsx')
print("Expenses sheet successfully updated. Open file to view changes.")
os.startfile('expenses.xlsx')
    


